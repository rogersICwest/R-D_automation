"""
supporting functions
"""
import pandas as pd
import numpy as np
import os
import re
import openpyxl # for modifying xlsx files
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Color, PatternFill, Font, Border
import xlrd # adding support for older xls files
import pyautogui

# if file is the older version xls, have to convert with this function first
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException
def open_xls_as_xlsx(filename, index):  # reading xls into openpyxl workbook
    # first open using xlrd
    book = xlrd.open_workbook(filename)
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.active

    for row in range(0, nrows):
        for col in range(0, ncols):
            sheet1.cell(row=row+1, column=col+1).value = sheet.cell_value(row, col)

    return book1

def read_xls_xlsx(filename, sheets_number):
    if filename[-4:] == "xlsx":
        workbook = openpyxl.load_workbook(filename)
        ws = workbook[workbook.sheetnames[sheets_number]]
    elif filename[-4:] == ".xls":
        workbook = open_xls_as_xlsx(filename, sheets_number)
        # because in this case it is converted, it has only the first sheet
        ws = workbook[workbook.sheetnames[0]]
    return ws

def read_val_into_set(filename):
    vals = set()
    try:
        ws = read_xls_xlsx(filename, 0)
    except:
        ws = read_xls_xlsx(filename + "x", 0)

    for i in range(ws.max_row):
        count_space = 0
        for j in range(99):
            val = ws.cell(row=i+1, column=j+1).value
            if val == "" or val is None: count_space += 1
            if count_space >= 5: break
            vals.add(val)
    
    return vals

def get_header(worksheet):
    # assume first row is the header
    # openpyxl is 1-indexed
    header = []
    for i in range(999):
        val = worksheet.cell(row=1, column=i+1).value
        if val == "" or val is None: break
        header.append(val)
    header = [x.lower() for x in header]
    # debug
    print("get header worked okay")
    return header

def get_cell_in_partNumber(worksheet, Partnumber_col):
    cells_in_partNumber = []
    for i in range(2,worksheet.max_row+1):
        val = worksheet.cell(row=i, column=Partnumber_col).value
        cells_in_partNumber.append(val)
    # debug
    print("cells in partnumber column collected")
    return cells_in_partNumber, set(cells_in_partNumber)


def label_rows(main_ws, Partnumber_col): # , folder
    # assume that in partnumber column, duplicate part numbers stay together
    # i.e. p1,p1,p1,p2,p2,p3,p3 instead of p1,p1,p2,p1,p2,p3...
    parts = get_cell_in_partNumber(main_ws, Partnumber_col)

    # define color fills
    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')
    greenFill = PatternFill(start_color='FF00FF00',
                          end_color='FF00FF00',
                          fill_type='solid')
    yellowFill = PatternFill(start_color='FFFFFF00',
                          end_color='FFFFFF00',
                          fill_type='solid')
    greyFill = PatternFill(start_color='FF808080',
                          end_color='FFFFFF00',
                          fill_type='solid')
    
    # loop through parts
    N = len(parts[1])
    for i, p in enumerate(parts[1]):
        row_ind = parts[0].index(p) + 2
        # debug
        # print("p here is: " + p)
        # print("row_ind here is: " + str(row_ind))
        progress = str(int(round(100*i/N, 0)))
        print("progress: " + progress + " percent completed")
        try:
            val_set = read_val_into_set(p + ".xls") # folder +
            # debug
            print(val_set)
        except:
            # if no file found, label grey
            while main_ws.cell(row=row_ind, column=Partnumber_col).value == p:
                main_ws.cell(row=row_ind, column=Partnumber_col).fill = greyFill
                row_ind += 1
            continue
        count_correct = 0
        # debug
        # print("val set length here is: " + str(len(val_set)))
        # loop through duplicates
        while main_ws.cell(row=row_ind, column=Partnumber_col).value == p:
            # debug
            # print("inside while: " + str(count_correct))
            incorrect = 0
            for i in range(1, Partnumber_col - 5):
                # need to adapt to both float and string type
                value2look = round(main_ws.cell(row=row_ind, column=i).value,5)
                if value2look not in val_set and str(value2look) not in val_set:
                    # debug 0305
                    main_ws.cell(row=row_ind, column=i).fill = yellowFill
                    # log # of incorrect
                    incorrect += 1
                else:
                    # debug 0305
                    main_ws.cell(row=row_ind, column=i).fill = greenFill
            if incorrect / (Partnumber_col - 5) < 0.2:
                if count_correct == 0:
                    main_ws.cell(row=row_ind, column=Partnumber_col).fill = greenFill
                    count_correct += 1
                else:
                    # TODO: go back and change the other to yellow as well
                    main_ws.cell(row=row_ind, column=Partnumber_col).fill = yellowFill
            else:
                main_ws.cell(row=row_ind, column=Partnumber_col).fill = redFill
            row_ind += 1

def find_file_helper(filelist, input_name):
    if input_name in filelist:
        return input_name
    elif (input_name + ".xls") in filelist:
        return input_name + ".xls"
    elif (input_name + ".xlsx") in filelist:
        return input_name + ".xlsx"
    else:
        return None

def duplicate_remover(): # folder, main_xlsx, sheets_number = 1
    main_xlsx = pyautogui.prompt('the main xlsx file')
    filelist = os.listdir()
    for _ in range(2):
        main_xlsx = find_file_helper(filelist, main_xlsx)
        if main_xlsx is None:
            main_xlsx = pyautogui.prompt('file not found, please retype the name')
    sheets_number = int(pyautogui.prompt('please specify which sheet to use'))
    sheets_number -= 1
    if main_xlsx[-4:] == "xlsx":
        try:
            main_wb = openpyxl.load_workbook(main_xlsx) # folder +
        except:
            pyautogui.alert('Error: File not found, are we looking at the right folder?')
        main_ws = main_wb[main_wb.sheetnames[sheets_number]]
        # debug
        print("xlsx file imported")
    elif main_xlsx[-4:] == ".xls":
        try:
            main_wb = open_xls_as_xlsx(main_xlsx, sheets_number) # folder +
        except:
            pyautogui.alert('Error: File not found, are we looking at the right folder?')
        # because in this case it is converted, it has only the first sheet
        main_ws = main_wb[main_wb.sheetnames[0]]
        # debug
        print("xls file imported")

    header = get_header(main_ws)
    Partnumber_col = header.index("partnumber") + 1
    # debug
    # print(Partnumber_col)
    # debug
    # parts = get_cell_in_partNumber(main_ws, Partnumber_col)
    # print(parts)
    label_rows(main_ws, Partnumber_col) #, folder

    '''
    option 1: name output file with user input
    '''
    # saveTo = pyautogui.prompt('process completed, please name the output file')
    # if .xls in saveTo:
    #     saveTo = saveTo.split(".")[0] + "xlsx"
    # else:
    #     saveTo += "xlsx"
    # main_wb.save(saveTo)
    # command = "start excel " + saveTo

    '''
    option 2: name output file with date
    '''
    # import datetime
    # currentDT = datetime.datetime.now()
    # saveTo = "labelled_result_" + str(currentDT.month) + "_" + str(currentDT.day) + ".xlsx"
    # main_wb.save(saveTo)
    # command = "start excel " + saveTo

    '''
    option 3: name output file with fixed name
    '''
    main_wb.save('labelled_result.xlsx')
    command = 'start excel labelled_result.xlsx'
    os.system(command)